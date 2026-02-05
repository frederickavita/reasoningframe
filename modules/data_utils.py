# =============================================================================
# MODULE: data_utils.py (VERSION V1.7 - STRICT PARSER & STREAMING)
# =============================================================================
import csv
import os
import re
import datetime

# TENTATIVE D'IMPORT DU MOTEUR "CLEVER"
try:
    import clevercsv
    HAS_CLEVER = True
except ImportError:
    HAS_CLEVER = False

# ✅ LISTE BLANCHE DES SÉPARATEURS AUTORISÉS (Pour éviter les faux positifs comme @)
VALID_DELIMITERS = {",", ";", "\t", "|"}

# ==============================================================================
# 1. HELPERS META / NORMALISATION
# ==============================================================================

def detect_encoding(file_path):
    """Détecte l'encodage (UTF-8, Latin-1, etc.)"""
    encodings_to_try = ["utf-8-sig", "utf-8", "cp1252", "latin-1", "iso-8859-1"]
    for enc in encodings_to_try:
        try:
            with open(file_path, "r", encoding=enc, errors="strict") as f:
                f.read(4096)
                return enc
        except UnicodeDecodeError:
            continue
    return "latin-1"

def detect_csv_dialect_smart(file_path, encoding):
    """
    Détecte le délimiteur mais REFUSE les délires (ex: @, ., :)
    """
    delimiter = "," # Par défaut
    quotechar = '"'
    
    try:
        with open(file_path, "r", encoding=encoding, errors="replace", newline='') as f:
            sample = f.read(16000)
            
            # 1. Essai CleverCSV
            if HAS_CLEVER:
                try:
                    d = clevercsv.Sniffer().sniff(sample, verbose=False)
                    if d.delimiter in VALID_DELIMITERS:
                        return d.delimiter, getattr(d, 'quotechar', '"')
                except:
                    pass
            
            # 2. Essai Standard
            try:
                sniffer = csv.Sniffer()
                d2 = sniffer.sniff(sample, delimiters=[',', ';', '\t', '|'])
                if d2.delimiter in VALID_DELIMITERS:
                    return d2.delimiter, getattr(d2, 'quotechar', '"')
            except:
                pass
            
            # 3. Fallback "Vote" (Si les sniffers échouent)
            counts = {sep: sample.count(sep) for sep in VALID_DELIMITERS}
            best = max(counts, key=counts.get) if counts else ","
            if counts.get(best, 0) > 0:
                delimiter = best

    except Exception:
        pass
        
    return delimiter, quotechar

def normalize_text_surface(text):
    if text is None: return ""
    text = str(text)
    return text.replace("\ufeff", "").replace("\xa0", " ").strip()

def to_snake_case(text):
    text = normalize_text_surface(text).lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text

def normalize_headers(raw_headers):
    cols = []
    seen = {}
    for idx, raw in enumerate(raw_headers):
        raw_str = normalize_text_surface(raw)
        norm = to_snake_case(raw_str) or f"column_{idx+1}"
        
        if norm in seen:
            seen[norm] += 1
            norm = f"{norm}_{seen[norm]}"
        else:
            seen[norm] = 1
            
        cols.append({"id": idx, "raw": raw_str, "norm": norm})
    return cols

# ==============================================================================
# 2. HEURISTIQUE HEADER
# ==============================================================================

def detect_header_row_index(sample_rows, lookahead=5):
    if not sample_rows: return 0

    limit = min(len(sample_rows), 20)
    max_cols = 0
    for r in sample_rows[:limit]:
        if r and any(str(x).strip() for x in r):
            max_cols = max(max_cols, len(r))
    if max_cols == 0: return 0

    def is_text_like(s):
        s = normalize_text_surface(s)
        if not s: return False
        if re.match(r'^[+-]?\d+([.,]\d+)?$', s): return False
        if re.match(r'^\d{2,4}[-/]\d{2}[-/]\d{2,4}', s): return False
        return True

    best_idx = 0
    best_score = -10**9

    for i in range(limit):
        row = sample_rows[i] or []
        width = len(row)
        cells = [normalize_text_surface(x) for x in row]
        non_empty = [c for c in cells if c]
        if not non_empty: continue

        uniq_ratio = len(set(non_empty)) / max(1, len(non_empty))
        text_ratio = sum(1 for c in non_empty if is_text_like(c)) / max(1, len(non_empty))
        fill_ratio = len(non_empty) / max(1, width)

        score = 2.0 * uniq_ratio + 3.0 * text_ratio + 1.0 * fill_ratio

        if max_cols >= 3 and width < int(max_cols * 0.6): score -= 10.0
        if width == max_cols: score += 1.0
        if i == 0 and score > -5: score += 0.5

        if score > best_score:
            best_score = score
            best_idx = i

    return best_idx

# ==============================================================================
# 3. TYPAGE SOFT (SECURISE)
# ==============================================================================

def infer_value_soft(raw_val):
    if raw_val is None: return None, None
    s = str(raw_val).strip()
    if s == "": return None, None
    
    # 1. Null Tokens
    if s.upper() in ("NULL", "N/A", "NAN", "NONE") or s in ("-", "—", "–"):
        return None, "NULL_TOKEN"
    
    # 2. Security (Injection)
    if s.startswith("=") or s.startswith("@"): 
        return s, "CSV_FORMULA_RISK"
    
    if s.startswith("+") or s.startswith("-"):
        if not re.match(r"^[+-]?\d+([.,]\d+)?$", s):
            return s, "CSV_FORMULA_RISK"
    
    # 3. Leading Zeros
    if s.startswith("0") and len(s) > 1 and not s.startswith("0.") and not s.startswith("0,"):
        return s, None

    # 4. Int
    if re.match(r"^-?\d+$", s):
        if len(s) < 15: return int(s), None
        return s, None

    # 5. Float
    if re.match(r"^[+-]?\d+[.,]\d+$", s):
        try: return float(s.replace(",", ".")), None
        except: pass
            
    return s, None

# ==============================================================================
# 4. MAIN READ FUNCTION
# ==============================================================================

def read_file_to_virtual_table(file_path, preview_limit=10):
    try:
        filename = os.path.basename(file_path).lower()

        # --- EXCEL ---
        if filename.endswith(".xlsx"):
            try:
                from openpyxl import load_workbook
                wb = load_workbook(file_path, read_only=True, data_only=True)
                ws = wb.active
                row_iter = ws.iter_rows(values_only=True)

                scan_buffer_size = 30
                raw_rows_buffer = []
                for _ in range(scan_buffer_size):
                    try:
                        r = next(row_iter)
                        raw_rows_buffer.append([str(c) if c is not None else "" for c in r])
                    except StopIteration: break

                if not raw_rows_buffer: return {"status": "error", "message": "Fichier Excel vide."}

                header_idx = detect_header_row_index(raw_rows_buffer)
                raw_headers = raw_rows_buffer[header_idx]
                columns = normalize_headers(raw_headers)
                col_count = len(columns)
                preview_rows = []
                total_rows = 0
                buffer_data = raw_rows_buffer[header_idx + 1:]

                def chain_rows():
                    for r in buffer_data: yield r
                    for r in row_iter: yield [str(c) if c is not None else "" for c in r]

                for row in chain_rows():
                    cells = [normalize_text_surface(c) for c in row[:col_count]]
                    while len(cells) < col_count: cells.append("")
                    if all(c == "" for c in cells): continue

                    total_rows += 1
                    if total_rows <= preview_limit:
                        typed = []
                        issues = {}
                        for i, raw_val in enumerate(cells):
                            v, issue = infer_value_soft(raw_val)
                            typed.append(v)
                            if issue: issues[columns[i]["norm"]] = issue
                        preview_rows.append({"_id": total_rows, "raw": cells, "typed": typed, "issues": issues})

                return {
                    "status": "success",
                    "file": {"name": os.path.basename(file_path), "type": "xlsx", "engine": "openpyxl", "detected_header_row": header_idx + 1},
                    "columns": columns, "rows": preview_rows, "total_rows": total_rows, "preview_limit": preview_limit
                }
            except ImportError:
                return {"status": "error", "message": "Module openpyxl manquant."}

        # --- CSV ---
        encoding = detect_encoding(file_path)
        delim, quote = detect_csv_dialect_smart(file_path, encoding)
        
        # SÉCURITÉ : Fallback virgule si échec détection
        if delim not in VALID_DELIMITERS:
            delim = ","

        scan_buffer_size = 30
        raw_rows_buffer = []

        with open(file_path, "r", encoding=encoding, newline="", errors="replace") as f:
            if HAS_CLEVER:
                reader = clevercsv.reader(f, delimiter=delim, quotechar=quote)
            else:
                reader = csv.reader(f, delimiter=delim, quotechar=quote)

            for _ in range(scan_buffer_size):
                try: raw_rows_buffer.append(next(reader))
                except StopIteration: break
                except Exception: break

            if not raw_rows_buffer: return {"status": "error", "message": "Fichier vide ou illisible."}

            header_idx = detect_header_row_index(raw_rows_buffer)
            raw_headers = raw_rows_buffer[header_idx]
            columns = normalize_headers(raw_headers)
            col_count = len(columns)
            preview_rows = []
            total_rows = 0
            buffer_data = raw_rows_buffer[header_idx + 1:]

            def chain_rows():
                for r in buffer_data: yield r
                for r in reader: yield r

            for row in chain_rows():
                row = row or []
                cells = [normalize_text_surface(c) for c in row[:col_count]]
                while len(cells) < col_count: cells.append("")
                if all(c == "" for c in cells): continue

                total_rows += 1
                if total_rows <= preview_limit:
                    typed = []
                    issues = {}
                    for i, raw_val in enumerate(cells):
                        v, issue = infer_value_soft(raw_val)
                        typed.append(v)
                        if issue: issues[columns[i]["norm"]] = issue
                    preview_rows.append({"_id": total_rows, "raw": cells, "typed": typed, "issues": issues})

        return {
            "status": "success",
            "file": {
                "name": os.path.basename(file_path), 
                "type": "csv", 
                "encoding": encoding, 
                "delimiter": delim, 
                "quotechar": quote, 
                "engine": "clevercsv" if HAS_CLEVER else "csv", 
                "detected_header_row": header_idx + 1
            },
            "columns": columns, "rows": preview_rows, "total_rows": total_rows, "preview_limit": preview_limit
        }

    except Exception as e:
        return {"status": "error", "message": f"Erreur Parser: {str(e)[:200]}"}