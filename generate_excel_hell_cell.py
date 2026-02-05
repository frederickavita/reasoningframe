from pathlib import Path
from openpyxl import Workbook

OUT = Path("test_datasets")
OUT.mkdir(parents=True, exist_ok=True)

nbsp = "\u00A0"
tab = "\t"
newline = "\n"

# Cellule "hell"
hell = (
    'Hello "Quote" and ""double quotes""'
    + f"{newline}Line2 after newline"
    + f"{newline}Here is a TAB:{tab}[TAB]"
    + f"{newline}Here is NBSP:59{nbsp}000"
    + f"{newline}Emoji: 🚀  Chinese: 淘宝"
)

wb = Workbook()
ws = wb.active
ws.title = "HellCell"

# Headers
ws.append(["id", "name", "notes", "status"])

# Row 1 (hell cell in notes)
ws.append([1, "Alice", hell, "OK"])

# Row 2 (normal row after hell)
ws.append([2, "Bob", "Normal row after hell cell", "OK"])

# Row 3 (empty row in the middle)
ws.append([None, None, None, None])

# Row 4 (normal row)
ws.append([3, "Charlie", "Another normal row", "OK"])

# Optional: make the hell cell wrap text in Excel viewers
ws["C2"].alignment = ws["C2"].alignment.copy(wrap_text=True)

path = OUT / "xlsx_unicode_hell_cell.xlsx"
wb.save(path)

print("✅ Generated:", path.resolve())
print("Hell cell repr:", repr(hell))
